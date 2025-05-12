from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'supersecretkey'

EXCEL_FILE = 'login_data.xlsx'

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Logins'
        ws.append(['Username', 'Password', 'Login Time', 'Logout Time', 'Project', 'Hours', 'Total Hours'])
        wb.save(EXCEL_FILE)

initialize_excel()

@app.route('/')
def home():
    return render_template('index22.html')

@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']
    login_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    session['username'] = username
    session['password'] = password
    session['login_time'] = login_time

    wb = load_workbook(EXCEL_FILE)
    ws = wb['Logins']
    ws.append([username, password, login_time, '', '', '', ''])
    wb.save(EXCEL_FILE)

    return redirect(url_for('project'))

@app.route('/project')
def project():
    return render_template('project22.html')

@app.route('/submit_project', methods=['POST'])
def submit_project():
    project = request.form['projects']
    hours = request.form['hours']
    session['project'] = project
    session['hours'] = hours
    return redirect(url_for('project'))

@app.route('/logout', methods=['POST'])
def logout():
    username = session.get('username')
    password = session.get('password')
    login_time = session.get('login_time')
    project = session.get('project')
    hours = session.get('hours')
    logout_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    login_dt = datetime.strptime(login_time, '%Y-%m-%d %H:%M:%S')
    logout_dt = datetime.strptime(logout_time, '%Y-%m-%d %H:%M:%S')
    total_hours = (logout_dt - login_dt).total_seconds() / 3600

    wb = load_workbook(EXCEL_FILE)

    if username not in wb.sheetnames:
        ws = wb.create_sheet(title=username)
        ws.append(['User ID', 'Password','Login Time', 'Logout Time', 'Project', 'Hours' ,'hours log'])
        ws.append([username, password, login_time, logout_time,project,hours,total_hours])
    else:
        ws = wb[username]
        ws.append([username, password, login_time, logout_time,project,hours,total_hours])
    

    wb.save(EXCEL_FILE)

    session.clear()
    return redirect(url_for('home'))

if __name__ == '__main__':
    app.run(debug=True)
